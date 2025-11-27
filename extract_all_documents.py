#!/usr/bin/env python3
"""
Script pour extraire le texte de tous les documents et les indexer dans Justicia RAG
"""

import os
import sys
from pathlib import Path

# Import des bibliothèques nécessaires
try:
    from docx import Document
    from openpyxl import load_workbook
    import PyPDF2
except ImportError:
    print("Installation des dépendances...")
    os.system("pip3 install python-docx openpyxl PyPDF2 -q")
    from docx import Document
    from openpyxl import load_workbook
    import PyPDF2

def extract_text_from_pdf(pdf_path):
    """Extrait le texte d'un PDF"""
    try:
        text = ""
        with open(pdf_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page in pdf_reader.pages:
                text += page.extract_text() + "\n"
        return text
    except Exception as e:
        print(f"Erreur PDF {pdf_path}: {e}")
        return ""

def extract_text_from_docx(docx_path):
    """Extrait le texte d'un DOCX"""
    try:
        doc = Document(docx_path)
        text = "\n".join([paragraph.text for paragraph in doc.paragraphs])
        return text
    except Exception as e:
        print(f"Erreur DOCX {docx_path}: {e}")
        return ""

def extract_text_from_xlsx(xlsx_path):
    """Extrait le texte d'un XLSX"""
    try:
        wb = load_workbook(xlsx_path, data_only=True)
        text = ""
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            text += f"\n=== Feuille: {sheet_name} ===\n"
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    text += row_text + "\n"
        return text
    except Exception as e:
        print(f"Erreur XLSX {xlsx_path}: {e}")
        return ""

def extract_document(file_path, output_dir):
    """Extrait le texte d'un document et le sauvegarde"""
    file_path = Path(file_path)
    if not file_path.exists():
        print(f"❌ Fichier introuvable: {file_path}")
        return None
    
    # Déterminer le type de fichier
    ext = file_path.suffix.lower()
    
    print(f"📄 Extraction: {file_path.name}")
    
    # Extraire selon le type
    if ext == '.pdf':
        text = extract_text_from_pdf(file_path)
    elif ext == '.docx':
        text = extract_text_from_docx(file_path)
    elif ext == '.xlsx':
        text = extract_text_from_xlsx(file_path)
    else:
        print(f"⚠️  Type non supporté: {ext}")
        return None
    
    if not text or len(text) < 10:
        print(f"⚠️  Texte vide ou trop court")
        return None
    
    # Sauvegarder le texte extrait
    output_file = output_dir / f"{file_path.stem}.txt"
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(text)
    
    char_count = len(text)
    print(f"✅ {char_count:,} caractères extraits → {output_file.name}")
    
    return {
        'file_path': str(file_path),
        'output_file': str(output_file),
        'char_count': char_count,
        'status': 'success'
    }

def main():
    upload_dir = Path("/home/ubuntu/upload")
    output_dir = Path("/home/ubuntu/extracted_texts")
    output_dir.mkdir(exist_ok=True)
    
    # Liste tous les fichiers
    files = list(upload_dir.glob("*.pdf")) + list(upload_dir.glob("*.docx")) + list(upload_dir.glob("*.xlsx"))
    
    print(f"\n🚀 Extraction de {len(files)} documents...\n")
    
    results = []
    for file_path in sorted(files):
        result = extract_document(file_path, output_dir)
        if result:
            results.append(result)
    
    print(f"\n✅ {len(results)} documents extraits avec succès")
    print(f"📊 Total: {sum(r['char_count'] for r in results):,} caractères")
    
    # Sauvegarder le résumé
    summary_file = output_dir / "_extraction_summary.txt"
    with open(summary_file, 'w', encoding='utf-8') as f:
        f.write(f"Extraction de {len(results)} documents\n")
        f.write(f"Total: {sum(r['char_count'] for r in results):,} caractères\n\n")
        for r in results:
            f.write(f"{Path(r['file_path']).name}: {r['char_count']:,} caractères\n")
    
    print(f"📝 Résumé sauvegardé: {summary_file}")

if __name__ == "__main__":
    main()
