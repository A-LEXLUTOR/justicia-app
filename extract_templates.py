import json
import os
from pathlib import Path

# Installer les dépendances si nécessaire
try:
    import docx
except ImportError:
    os.system("pip3 install python-docx openpyxl --quiet")
    import docx

import openpyxl

def extract_docx(file_path):
    """Extraire le contenu d'un fichier DOCX"""
    try:
        doc = docx.Document(file_path)
        content = []
        for para in doc.paragraphs:
            if para.text.strip():
                content.append(para.text)
        return "\n".join(content)
    except Exception as e:
        return f"Erreur d'extraction: {str(e)}"

def extract_xlsx(file_path):
    """Extraire le contenu d'un fichier XLSX"""
    try:
        wb = openpyxl.load_workbook(file_path)
        content = []
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            content.append(f"=== {sheet_name} ===")
            for row in sheet.iter_rows(values_only=True):
                row_text = "\t".join([str(cell) if cell is not None else "" for cell in row])
                if row_text.strip():
                    content.append(row_text)
        return "\n".join(content)
    except Exception as e:
        return f"Erreur d'extraction: {str(e)}"

def main():
    upload_dir = Path("/home/ubuntu/upload")
    templates = []
    
    # Liste des fichiers
    files = sorted(upload_dir.glob("*.docx")) + sorted(upload_dir.glob("*.xlsx"))
    
    print(f"Extraction de {len(files)} modèles...")
    
    for i, file_path in enumerate(files, 1):
        print(f"[{i}/{len(files)}] {file_path.name}")
        
        # Extraire le contenu
        if file_path.suffix == ".docx":
            content = extract_docx(file_path)
        elif file_path.suffix == ".xlsx":
            content = extract_xlsx(file_path)
        else:
            continue
        
        # Créer l'entrée du modèle
        template = {
            "id": file_path.stem.lower().replace(" ", "-"),
            "name": file_path.stem.replace("PORTEO", "PORTEO ").replace("MODELE", "Modèle"),
            "filename": file_path.name,
            "type": file_path.suffix[1:],
            "content": content,
            "fields": []  # Sera rempli plus tard
        }
        
        templates.append(template)
    
    # Sauvegarder dans un fichier JSON
    output_file = "/home/ubuntu/data/templates_content.json"
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(templates, f, ensure_ascii=False, indent=2)
    
    print(f"\n✅ Extraction terminée ! {len(templates)} modèles sauvegardés dans {output_file}")

if __name__ == "__main__":
    main()
